import openpyxl
import random

data = openpyxl.load_workbook('events.xlsx')
event = data.worksheets[0]

for i in range(0, 1024):
	event.cell(i + 1, 1).value = random.randint(0, 1000)
	event.cell(i + 1, 2).value = str("event" + str(random.randint(0, 100)))

data.save('events.xlsx')