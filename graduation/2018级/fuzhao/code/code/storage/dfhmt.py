import random
import math
import hashlib
import openpyxl
import couchdb
import datetime

#连接数据库
couch = couchdb.Server('http://admin:123456@39.106.100.60:5984/')#可修改
db = couch['weight']
#进制转换
def hex2dec(string_num):#10转16
    return str(int(string_num.upper(), 16))
base = [str(x) for x in range(10)] + [ chr(x) for x in range(ord('A'),ord('A')+6)]

def dec2hex(string_num):#16转10
    num = int(string_num)
    mid = []
    while True:
        if num == 0: break
        num,rem = divmod(num, 16)
        mid.append(base[rem])

    return ''.join([str(x) for x in mid[::-1]])




def g_weight(i):#随机生成权重序列，树的高度为i，边数为(2**n - 2)
	weight = []
	for i in range(0, 2**i - 2):
		weight.append(random.randint(0, 100))
	return weight

n = 11#树的高度，可修改
weight = g_weight(n)#生成权重序列

#将event读到event中
def get_data():
	data = openpyxl.load_workbook('events.xlsx')
	d_event = data.worksheets[0]
	event = []
	num = d_event.max_row#数据中为1024个，根据树的高度不同可以改变
	for i in range(0, num):
		event.append(str(d_event.cell(i + 1, 1).value) + str(d_event.cell(i + 1, 2).value))

	return event
#得到事件的hash值列表
def get_hash(event):
	#event = get_data()
	e_hash = []
	for i in event:
		h = hashlib.sha256()
		h.update(str(i).encode('utf-8'))
		e_hash.append(h.hexdigest())
	#print(h.hexdigest())
	return e_hash


#根据e_hash, weight构建树，返回的node表示树的节点，从叶子节点向上记录，最后一个为树的根
def build_tree(weight):
	e_hash = get_hash(get_data())
	node = []
	node = node + e_hash
	leaf = e_hash
	index = 0
	
	for i in range(n - 2, -1, -1):
		parent = []
		for j in range(0, 2**i):
			#print(leaf[i * 2] ," ", weight[index] ," ", leaf[i * 2 + 1] ," ", weight[index + 1],"  ")
			left = int(hex2dec(str(leaf[j * 2])))
			right = int(hex2dec(str(leaf[j * 2 + 1])))
			p = dec2hex(str(left * weight[index] + right * weight[index + 1])).lower()
			#print(p)
			parent.append(str(p))
			#print(len(parent))
			index = index + 2
		leaf = parent
		#print(len(leaf))
		node = node + parent
	return node

tree_node = build_tree(weight)

def v_node(i, n):#返回验证路径
	node = [] #节点到根节点路径
	bro_node = [] #兄弟节点，即验证路径
	l = [0]
	index = 0
	#每层的节点数
	for j in range(1, n - 1):
		l.append(2 ** (n - j) + l[j - 1])

	for j in range(0, n - 1):
		index = l[j] + int(i / (2 ** j))
		node.append(index)

	for j in node:
		if j % 2 == 0:
			bro_node.append(j + 1)
		else:
			bro_node.append(j - 1)
	return bro_node

def r_node(i, n):#待验证节点到根节点的路径
	node = [] #节点到根节点路径
	l = [0]
	index = 0
	#每层的节点数
	for j in range(1, n - 1):
		l.append(2 ** (n - j) + l[j - 1])

	for j in range(0, n - 1):
		index = l[j] + int(i / (2 ** j))
		node.append(index)
	return node


#从数据库中查寻
def get_weight(i):#根据weight的id，查找weight
	doc = db['5d65a00ff1b85868fd26af22ef020165']#数据库中文件id
	a = doc[str(i)]
	return a

def get_root():#查找根节点
	doc = db['5d65a00ff1b85868fd26af22ef020165']#数据库中文件id
	a = doc['root']
	return a

def get_node(i):#查找服务器里存的验证路径
	doc = db['5d65a00ff1b85868fd26af22ef02035b']#数据库中文件id
	a = doc[str(i)]
	return a

l = len(tree_node)
#print(l)
root_node = tree_node[l - 1]
#print(root_node)
#print(tree_node[43])

#根据数据data即其在树中的位置验证是否正确
def verify(i, data):
	v_n = v_node(i, n)
	r_n = r_node(i, n)
	#
	r_child = str(data)
	#重新计算根节点
	for j in range(0, n - 1):
		v_child = int(hex2dec(str(get_node(v_n[j]))))
		r_child = int(hex2dec(str(r_child)))
		#print(v_n[j], ' ', r_n[j])
		#print(v_child, ' ', r_child)
		p = v_child * get_weight(v_n[j]) + r_child * get_weight(r_n[j])
		parent = dec2hex(str(p)).lower()
		r_child = parent
		root_node = get_root()
		#print(root_node)
	if parent == root_node:
		return 1#正确
	else:
		return 0#错误
'''

def verify(i, data):
	v_n = v_node(i, n)
	r_n = r_node(i, n)
	#
	r_child = str(data)
	for j in range(0, n - 1):
		v_child = int(hex2dec(str(tree_node[v_n[j]])))
		r_child = int(hex2dec(str(r_child)))
		#print(v_n[j], ' ', r_n[j])
		#print(v_child, ' ', r_child)
		p = v_child * weight[v_n[j]] + r_child * weight[r_n[j]]
		parent = dec2hex(str(p)).lower()
		r_child = parent
		#root_node = get_root()
		#print(root_node)
	if parent == root_node:
		return 1
	else:
		return 0
'''

#验证时间
'''
start = datetime.datetime.now()
print(verify(43, 'bfc080cf0ea58dc05cd5dfc7f7ec40e09687129ea6a9669047e9f23d94df49ca'))
end = datetime.datetime.now()
print(end - start)
'''

#weight和根节点存储到数据库
def weight_s(weight):
	
	w_dict = {}
	l = len(weight)
	
	for i in range(0, l):
		#print(i)
		w_dict[str(i)] = weight[i]

	l_n = len(tree_node)
	w_dict['root'] = tree_node[l_n - 1]
	#print(w_dict)
	db.save(w_dict)

#存储所有节点，模拟云服务器，根据其返回验证路径
def node_s(tree_node):
	ans = {}
	l = len(tree_node)
	for i in range(0, l):
		ans[str(i)] = tree_node[i]
	db.save(ans)
weight_s(weight)
#node_s(tree_node)



