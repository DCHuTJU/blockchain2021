import random
import math
import hashlib
import openpyxl
import couchdb

#直接存储hash值的方法
couch = couchdb.Server('http://127.0.0.1:5984/')
db = couch['weight']


def get_data():
	data = openpyxl.load_workbook('events.xlsx')
	d_event = data.worksheets[0]
	event = []
	num = d_event.max_row#数据中为1024个，根据树的高度不同可以改变
	#num = #可以调整数据的大小
	for i in range(0, num):#
		event.append(str(d_event.cell(i + 1, 1).value) + str(d_event.cell(i + 1, 2).value))
	return event

def get_hash(event):
	#event = get_data()
	e_hash = []
	for i in event:
		h = hashlib.sha256()
		h.update(str(i).encode('utf-8'))
		e_hash.append(h.hexdigest() + h.hexdigest())#dam中存证数据长度为512
	#print(h.hexdigest())
	return e_hash




def hash_s():
	event = get_data()
	h = get_hash(event)
	
	h_dict = {}
	l = len(h)
	
	for i in range(0, l):
		#print(i)
		h_dict[str(i)] = h[i]

	print(h_dict)
	#db.save(w_dict)

hash_s()