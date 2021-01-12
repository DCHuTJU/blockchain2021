import openpyxl
#读取数据打包成交易

def getoffsetlist():
	data = openpyxl.load_workbook("offset.xlsx")
	offset = data.worksheets[0]

	num = offset.max_row
	offsetlist = []
	#每个块128笔交易
	for i in range(0, 128):
		offsetlist.append([offset.cell(i + 1,1).value,offset.cell(i + 1,2).value])
	#print(offsetlist)
	return offsetlist

#getoffsetlist()