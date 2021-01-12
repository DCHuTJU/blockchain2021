#计算更新延迟
import openpyxl
import random

#读数据
def getoffsetlist():
	data = openpyxl.load_workbook("offset.xlsx")
	offset = data.worksheets[0]

	num = offset.max_row
	offsetlist = []#[a, b, c, d]代表车辆id，偏移量，生成时区块高度，记录进块时区块高度
	for i in range(0, num):
		offsetlist.append([offset.cell(i + 1,1).value,offset.cell(i + 1,2).value, 0, 0])
	#print(offsetlist)

	#加时间信息
	#初始时有256笔交易，每生成一个块后新增128笔交易
	for i in range(0, 256):
		offsetlist[i][2] = 0
	
	for i in range(256, num):
		offsetlist[i][2] = int(i / 128 - 1)

	#print(offsetlist)

	return offsetlist

#随机选择128个块打包
def choose_r():
	a = [0] * 256
	for i in range(0, 256):
		a[i] = i
	ans = random.sample(a,128)#选择128个打包成一个块
	return ans

#将选择的交易剔除
def deal_tr(offsetlist):
	temp = []
	for i in offsetlist:
		if i[3] == 0:
			temp.append(i)
	return temp

#将选择的交易打包
def pac_tr(offsetlist):
	temp = []
	for i in offsetlist:
		if i[3] != 0:
			temp.append(i)
	return temp

#多次随机选择，模拟多个rsu竞争，最后选择绝对值之和最大的一个块
def choose_max(offsetlist):
	max_offset = []
	all_offset = 0
	#print(offsetlist)
	for i in range(0, 100):
		choose = choose_r()
		#print(choose)
		temp = 0
		for j in choose:
			temp = temp + abs(offsetlist[j][1])
		#print(temp)
		if temp > all_offset:
			all_offset = temp
			max_offset = choose
	return max_offset



#一共生成25个块，为每个块标记记录进区块链的时间
def tr_pool(offsetlist):
	block = []#块信息
	#print(offsetlist)
	for i in range(0,25):
		choose = choose_max(offsetlist)
		#print(choose)
		#为交易标记轮次
		for a in choose:
			offsetlist[a][3] = i + 1
			#print(a, offsetlist[3])
		#print(offsetlist)
		block.append(pac_tr(offsetlist))#将选择的交易打包
		offsetlist = deal_tr(offsetlist)#将选择的交易剔除
	return block

#计算不同大小的交易的平均延迟
def co_delay(block):
	delay = [0,0,0,0,0,0,0,0,0,0]
	count = [0,0,0,0,0,0,0,0,0,0]

	for i in block:
		#print(i)
		for j in i:
			#print(delay[0],j[3])
			d = j[3] - j[2]
			if d > 5:
				d = 5

			if abs(j[1]) >= 0 and abs(j[1]) < 0.5:
				#print(j[2])
				delay[0] = delay[0] + d
				count[0] = count[0] + 1
			elif abs(j[1]) >= 0.5 and abs(j[1]) < 1:
				delay[1] = delay[1] + d
				count[1] = count[1] + 1
			elif abs(j[1]) >= 1 and abs(j[1]) < 1.5:
				delay[2] = delay[2] + d
				count[2] = count[2] + 1
			elif abs(j[1]) >= 1.5 and abs(j[1]) < 2:
				delay[3] = delay[3] + d
				count[3] = count[3] + 1
			elif abs(j[1]) >= 2 and abs(j[1]) < 2.5:
				delay[4] = delay[4] + d
				count[4] = count[4] + 1
			elif abs(j[1]) >= 2.5 and abs(j[1]) < 3:
				delay[5] = delay[5] + d
				count[5] = count[5] + 1
			elif abs(j[1]) >= 3 and abs(j[1]) < 3.5:
				delay[6] = delay[6] + d
				count[6] = count[6] + 1
			elif abs(j[1]) >= 3.5 and abs(j[1]) < 4:
				delay[7] = delay[7] + d
				count[7] = count[7] + 1
			elif abs(j[1]) >= 4 and abs(j[1]) < 4.5:
				delay[8] = delay[8] + d
				count[8] = count[8] + 1
			elif abs(j[1]) >= 4.5 and abs(j[1]) < 5:
				delay[9] = delay[9] + d
				count[9] = count[9] + 1

	ans = []
	#print(delay)
	#print(count)
	for i in range(0, 10):
		if count[i] != 0:
			avg_delay = delay[i] / count[i]
			ans.append(round(avg_delay, 4))
		else:
			ans.append(0)

	return ans


#计算每个块的延迟
def co_all_delay(block):
	d_cost = []
	for i in block:
		delay = 0
		for j in i:
			delay = delay + abs(j[1] * (j[3] - j[2]))
		d_cost.append(round(delay / 128, 4))
	return d_cost 



offsetlist = getoffsetlist()
block = tr_pool(offsetlist)
avg_delay = co_delay(block)
print(avg_delay)
#d_cost = co_all_delay(block)
#print(d_cost)
#计算所有块的累计延迟
def add_cost(d_cost):
	cost = []
	temp = 0
	for i in d_cost:
		temp = temp + i
		cost.append(round(temp, 4))
	return cost

#ans = add_cost(d_cost)
#print(ans)
