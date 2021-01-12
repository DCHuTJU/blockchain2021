#信誉值随轮次变化，第6轮和第15轮报告错误信息，从0开始
#0.162; 0.279; 0.276; 0.167; 0.1; 0.096; 0.084对应0%，5%，10%，15%，20%，25%，30%
import math
import openpyxl
import random
import random_g

data = openpyxl.load_workbook('car-rsu.xlsx')
car_name = data.worksheets[0]
#行数
n = car_name.max_row #有多少个rsu

rsu_car = [0] * n #每个rsu有多少辆车

car_num = 0

#print(car_name.cell(17,1).value)

for i in range(1, n + 1):
	temp = car_name.cell(i,1).value
	rsu_car[i - 1] = temp
	#print(type(temp))
	car_num = car_num + temp
#print(car_num)

car = []#信誉值，数据可信度，是否正确，类别
#car = [a] * car_num
for i in range(0, car_num):
	car.append([0]*4)

def deal_data():#为每辆车赋值（是否正确，可信度->信誉值）
	per = [0.139, 0.256, 0.253, 0.144, 0.077, 0.073, 0.058]#不同的比例
	car_per = []
	temp = 0
	for i in range(0, 6):
		temp = temp + int(per[i] * car_num)
		car_per.append([temp, i])
	car_per.append([car_num, 6])#不同等级的分布
	#print(car_per)
	v = random_g.random_nr(car_num, car_num)#车辆id随机排序
	#print(mv)

	#为每辆车分配等级
	for i in v:
		for j in car_per:
			if i > j[0] - 1:
				car[i][3] = j[1]

	#将恶意车辆置为1，诚实车辆置为0
	#根据车辆类别有不同概率作恶
	for i in car:
		if i[3] == 0:
			i[2] = 0
		elif i[3] == 1:
			if random.random() < 0.05:
				i[2] = 1
			else:
				i[2] = 0
		elif i[3] == 2:
			if random.random() < 0.1:
				i[2] = 1
			else:
				i[2] = 0
		elif i[3] == 3:
			if random.random() < 0.15:
				i[2] = 1
			else:
				i[2] = 0
		elif i[3] == 4:
			if random.random() < 0.2:
				i[2] = 1
			else:
				i[2] = 0
		elif i[3] == 5:
			if random.random() < 0.25:
				i[2] = 1
			else:
				i[2] = 0
		elif i[3] == 6:
			if random.random() < 0.3:
				i[2] = 1
			else:
				i[2] = 0
	
	#为车辆分配信息可信度
	for i in range(0, car_num):
		if car[i][2] == 0:
			car[i][1] = round(random.uniform(0.7, 0.9),2)
		else:
			car[i][1] = round(random.uniform(0.5, 0.8),2)

	#车辆当前信誉值
	for i in range(0, car_num):
		car[i][0] = int((math.exp(1 / math.log(1 / (car[i][1] - 0.1))) - 1) * 10)


def rsu_de():#把车辆分给不同的rsu
	left = 0#起始位置
	car_data = []
	for i in range(0, n):#n个rsu
		car_n = rsu_car[i] #这个rsu的车辆数
		right = left + car_n - 1#中止位置
		car_r = []
		for i in range(0, car_n):
			car_r.append([0]*3)

		for j in range(0, car_n):
			car_r[j][0] = car[left + j][0]
			car_r[j][1] = car[left + j][1]
			car_r[j][2] = car[left + j][2]
		left = right + 1
		car_data.append(car_r)
	#print(car_data[0])
	#print(car_data[n - 1])
	return(car_data)


def rsu_bys(car): #单个rsu计算贝叶斯
	e = 0.999 #先验概率
	p = 1
	pe = 1
	for i in car:
		if i[2] == 0:#是正确事件
			p = p * i[1]
			pe = pe * (1 - i[1])
			#print(i[1])
		else:#是错误事件
			p = p *(1 - i[1])
			pe = pe * i[1]
			#print(i[1])
	#print(e * p, (1 - e) * pe)
	cr = (e * p) / (e * p + ((1 - e) * pe))
	#print(cr, crx)
	return cr

#计算每个rsu得到的可信度
def rsu_all(d, celi):
	for i in range(0, n):
		celi[i] = round(rsu_bys(d[i]),4)
	#print(celi)
	return celi


def event(celi):
	t = 0
	f = 0
	for i in range(0, n):
		if celi[i] >= 0.5:
			t = t + 1
		else:
			f = f + 1
	if t > f:
		return 0, t, f#事件可信
	else:
		return 1, t, f #事件不可信



	


def change_repu():
	repu = [5]
	re = 5
	offset = 0
	for i in range(0, 25):
		deal_data()
		d = rsu_de()

		celi = []
		for a in range(0, n):
			celi.append([0])

		rsu_all(d, celi)
		
		#res = event(celi)[0]
		t = event(celi)[1]
		f = event(celi)[2]
		
		if i == 5:
			offset = round(-1 * abs(f - t)/ (t + f),4) * 0.8
		elif i == 14:
			offset = round(-1 * abs(f - t) / (t + f) ,4) * 0.8
		else:
			offset = round(abs(t - f) / (t + f) ,4) * 0.8

		#print(offset)


		re = re + offset
		repu.append(round(re,4))
	return(repu)

#change_repu()
print(change_repu())