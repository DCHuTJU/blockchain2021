import math
import openpyxl
import random
import random_g
data = openpyxl.load_workbook('car-rsu.xlsx')
car_name = data.worksheets[0]
#行数
n = car_name.max_row #有多少个rsu

rsu_car = [0] * n #每个rsu有多少辆车

car_num = 0

#print(car_name.cell(17,1).value)

for i in range(1, n + 1):
	temp = car_name.cell(i,1).value
	rsu_car[i - 1] = temp
	#print(type(temp))
	car_num = car_num + temp
#print(rsu_car)

car = []#信誉值，数据可信度，是否正确
#car = [a] * car_num
for i in range(0, car_num):
	car.append([0]*3)

#print(car[798][2])

#print(car_name.cell(1,1).value)
#car[0][2] = 1
#print(car)


def deal_data(a):#为每辆车赋值（是否正确，可信度->信誉值）
	mv_per = 0.55#恶意车辆比例
	mv_num = int(mv_per * car_num) #恶意车辆数目
	tv_num = car_num - mv_num #诚实车辆数目
	#print(car_num, mv_num)
	mv = random_g.random_nr(int(car_num), int(mv_num))#恶意车辆的id
	#print(mv)
	#将恶意车辆置为1
	#print(car)
	for i in range(0, mv_num):
		#print(i," ")
		car[mv[i]][2] = 1
		#print(i,' ', mv[i], " ")
	#print(car)
	
	#为车辆分配信息可信度
	for i in range(0, car_num):
		if car[i][2] == 0:
			car[i][1] = round(random.uniform(0.7, 0.9),2)
		else:
			car[i][1] = round(random.uniform(0.5, 0.8),2)

	#车辆当前信誉值
	for i in range(0, car_num):
		car[i][0] = int((math.exp(1 / math.log(1 / (car[i][1] - 0.1))) - 1) * 10)
	#print(car)


#根据车辆的信誉值计算事件的可信度，用不到了
#def rup_cr(i):
#	c = math.exp(-1 / log(1 + i))
#	return c

def rsu_de(car):#把车辆分给不同的rsu
	left = 0#起始位置
	car_data = []
	for i in range(0, n):#n个rsu
		car_n = rsu_car[i] #这个rsu的车辆数
		right = left + car_n - 1#中止位置
		car_r = []
		for i in range(0, car_n):
			car_r.append([0]*3)

		for j in range(0, car_n):
			car_r[j][0] = car[left + j][0]
			car_r[j][1] = car[left + j][1]
			car_r[j][2] = car[left + j][2]
		left = right + 1
		car_data.append(car_r)
	#print(car_data[0])
	#print(car_data[n - 1])
	return(car_data)

deal_data(car)
d = rsu_de(car)


def rsu_wv(car): #单个rsu计算贝叶斯
	t = 0
	f = 0
	for i in car:
		if i[2] == 0:
			t = t + i[1]
		else:
			f = f + i[1]
	if t > f:
		return 1
	else:
		return 0;


celi = []
for i in range(0, n):
	celi.append([0])


def rsu_all(n):
	for i in range(0, n):
		celi[i] = round(rsu_wv(d[i]),4)
	#print(celi)
	return celi
	

rsu_all(n)

def event(celi):
	t = 0;
	f = 0;
	for i in range(0, n):
		if celi[i] >= 0.5:
			t = t + 1
		else:
			f = f + 1
	if t > f:
		return 0
	else:
		return 1

print(event(celi))



#print(d[0])
#print(rus_bys(d[0]))




